from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.common.models import Region

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


REGION_LEVELS = {
    "регион",
    "город федерального значения",
}


class Command(BaseCommand):
    help = "Sync full Russian region names from the allsettlements Excel dataset."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            default=str(
                Path(settings.BASE_DIR)
                / "docs"
                / "data_allsettlements_anon_156_v20251217.xlsx"
            ),
            help="Path to data_allsettlements_anon_156_v20251217.xlsx.",
        )
        parser.add_argument(
            "--sheet",
            default="data",
            help="Workbook sheet name with region data.",
        )
        parser.add_argument(
            "--country-code",
            default="RU",
            help="Country code to write into Region rows.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Print stats without writing to the database.",
        )

    def handle(self, *args, **options):
        if load_workbook is None:
            raise CommandError("openpyxl is required to import XLSX files.")

        dataset_path = Path(options["file"])
        sheet_name = options["sheet"]
        country_code = options["country_code"].strip().upper()
        dry_run = options["dry_run"]

        self._validate_file(dataset_path)
        full_region_names = self._load_full_region_names(dataset_path, sheet_name)

        stats = {
            "created": 0,
            "renamed": 0,
            "unchanged": 0,
            "skipped_conflict": 0,
        }

        with transaction.atomic():
            for full_name in full_region_names:
                self._sync_region(full_name, country_code, dry_run, stats)
            if dry_run:
                transaction.set_rollback(True)

        prefix = (
            "Allsettlements regions dry run completed"
            if dry_run
            else "Allsettlements regions sync completed"
        )
        self.stdout.write(
            self.style.SUCCESS(
                f"{prefix}: "
                f"created={stats['created']}, "
                f"renamed={stats['renamed']}, "
                f"unchanged={stats['unchanged']}, "
                f"skipped_conflict={stats['skipped_conflict']}."
            )
        )

    def _validate_file(self, path):
        if not path.exists():
            raise CommandError(f"Dataset file does not exist: {path}")
        if not path.is_file():
            raise CommandError(f"Dataset path is not a file: {path}")

    def _load_full_region_names(self, path, sheet_name):
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if sheet_name not in workbook.sheetnames:
                raise CommandError(f"Workbook sheet does not exist: {sheet_name}")
            worksheet = workbook[sheet_name]
            rows = worksheet.iter_rows(values_only=True)
            try:
                headers = next(rows)
            except StopIteration as exc:
                raise CommandError(f"Workbook sheet is empty: {sheet_name}") from exc

            header_indexes = {
                str(value or "").strip(): index for index, value in enumerate(headers)
            }
            self._validate_headers(header_indexes)

            names = []
            seen = set()
            for row in rows:
                level = self._cell(row, header_indexes, "object_level")
                if self._normalize_key(level) not in REGION_LEVELS:
                    continue

                full_name = self._normalize_name(
                    self._cell(row, header_indexes, "region")
                    or self._cell(row, header_indexes, "object_name")
                )
                if full_name == "Российская Федерация":
                    continue

                normalized = self._normalize_key(full_name)
                if full_name and normalized not in seen:
                    seen.add(normalized)
                    names.append(full_name)
            return names
        finally:
            workbook.close()

    def _validate_headers(self, header_indexes):
        required_headers = {"object_level", "object_name", "region"}
        missing_headers = sorted(required_headers - set(header_indexes))
        if missing_headers:
            raise CommandError(
                "Workbook does not contain required columns: "
                + ", ".join(missing_headers)
            )

    def _sync_region(self, full_name, country_code, dry_run, stats):
        exact_region = Region.objects.filter(
            name=full_name,
            country_code=country_code,
        ).first()
        if exact_region:
            stats["unchanged"] += 1
            return

        short_region = self._find_short_region(full_name, country_code)
        if short_region is None:
            if not dry_run:
                Region.objects.create(name=full_name, country_code=country_code)
            stats["created"] += 1
            return

        if Region.objects.filter(name=full_name, country_code=country_code).exists():
            stats["skipped_conflict"] += 1
            return

        if not dry_run:
            short_region.name = full_name
            short_region.save(update_fields=("name", "updated_at"))
        stats["renamed"] += 1

    def _find_short_region(self, full_name, country_code):
        full_aliases = self._region_aliases(full_name)
        candidates = Region.objects.filter(country_code=country_code)
        for region in candidates:
            if self._normalize_key(region.name) in full_aliases:
                return region
        return None

    def _region_aliases(self, name):
        normalized = self._normalize_key(name)
        aliases = {normalized}

        prefixes = (
            "республика ",
            "город федерального значения ",
        )
        suffixes = (
            " область",
            " край",
            " автономная область",
            " автономный округ",
        )

        for prefix in prefixes:
            if normalized.startswith(prefix):
                aliases.add(normalized.removeprefix(prefix))
        for suffix in suffixes:
            if normalized.endswith(suffix):
                aliases.add(normalized.removesuffix(suffix))

        replacements = {
            "республика саха (якутия)": "саха /якутия/",
            "республика северная осетия - алания": "северная осетия - алания",
            "республика северная осетия алания": "северная осетия - алания",
            "ханты-мансийский автономный округ югра": "ханты-мансийский",
            "ямало-ненецкий автономный округ": "ямало-ненецкий",
            "чукотский автономный округ": "чукотский",
            "ненецкий автономный округ": "ненецкий",
            "удмуртская республика": "удмуртская",
            "чувашская республика": "чувашская",
            "чеченская республика": "чеченская",
            "кабардино-балкарская республика": "кабардино-балкарская",
            "карачаево-черкесская республика": "карачаево-черкесская",
        }
        replacement = replacements.get(normalized)
        if replacement:
            aliases.add(replacement)

        return aliases

    def _cell(self, row, header_indexes, name):
        index = header_indexes[name]
        if index >= len(row):
            return ""
        value = row[index]
        return "" if value is None else value

    def _normalize_name(self, value):
        if value is None:
            return ""
        return str(value).strip()[:150]

    def _normalize_key(self, value):
        if value is None:
            return ""
        return (
            " ".join(str(value).strip().casefold().replace("ё", "е").split())
            .replace(" — ", " - ")
            .replace("—", "-")
        )
