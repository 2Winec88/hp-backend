from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.common.models import City, Region

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


SETTLEMENT_LEVELS = {
    "населенный пункт",
    "населённый пункт",
    "город федерального значения",
}
COORDINATE_QUANT = Decimal("0.000001")


class Command(BaseCommand):
    help = "Import Russian settlements from the allsettlements Excel dataset."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            default=str(
                Path(settings.BASE_DIR)
                / "docs"
                / "data_allsettlements_anon_156_v20251217.xlsx"
            ),
            help="Path to data_allsettlements_anon_156_v20251217.xlsx.",
        )
        parser.add_argument(
            "--sheet",
            default="data",
            help="Workbook sheet name with settlement data.",
        )
        parser.add_argument(
            "--country-code",
            default="RU",
            help="Country code to write into City rows.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse the file and print stats without writing to the database.",
        )
        parser.add_argument(
            "--limit",
            type=int,
            default=None,
            help="Import at most this many settlement rows. Useful for checks.",
        )

    def handle(self, *args, **options):
        if load_workbook is None:
            raise CommandError("openpyxl is required to import XLSX files.")

        dataset_path = Path(options["file"])
        sheet_name = options["sheet"]
        country_code = options["country_code"].strip().upper()
        dry_run = options["dry_run"]
        limit = options["limit"]

        self._validate_file(dataset_path)
        regions_by_name = self._load_regions(country_code)
        cities_by_key = self._load_cities(country_code)
        self._seen_city_keys = set()

        stats = {
            "created": 0,
            "updated": 0,
            "unchanged": 0,
            "skipped_not_settlement": 0,
            "skipped_without_name": 0,
            "skipped_without_region": 0,
            "skipped_duplicate_in_file": 0,
            "skipped_invalid_coordinates": 0,
        }

        with transaction.atomic():
            for settlement in self._iter_settlements(dataset_path, sheet_name):
                if limit is not None and (
                    stats["created"] + stats["updated"] + stats["unchanged"]
                ) >= limit:
                    break
                self._import_settlement(
                    settlement,
                    regions_by_name,
                    cities_by_key,
                    country_code,
                    dry_run,
                    stats,
                )
            if dry_run:
                transaction.set_rollback(True)

        prefix = (
            "Allsettlements dry run completed"
            if dry_run
            else "Allsettlements import completed"
        )
        self.stdout.write(
            self.style.SUCCESS(
                f"{prefix}: "
                f"created={stats['created']}, "
                f"updated={stats['updated']}, "
                f"unchanged={stats['unchanged']}, "
                f"skipped_not_settlement={stats['skipped_not_settlement']}, "
                f"skipped_without_name={stats['skipped_without_name']}, "
                f"skipped_without_region={stats['skipped_without_region']}, "
                f"skipped_duplicate_in_file={stats['skipped_duplicate_in_file']}, "
                f"skipped_invalid_coordinates={stats['skipped_invalid_coordinates']}."
            )
        )

    def _validate_file(self, path):
        if not path.exists():
            raise CommandError(f"Dataset file does not exist: {path}")
        if not path.is_file():
            raise CommandError(f"Dataset path is not a file: {path}")

    def _load_regions(self, country_code):
        regions = Region.objects.filter(country_code__in=(country_code, "")).order_by(
            "country_code"
        )
        return {self._normalize_key(region.name): region for region in regions}

    def _load_cities(self, country_code):
        cities = City.objects.select_related("region").filter(country_code=country_code)
        return {
            (self._normalize_key(city.name), city.region_id, country_code): city
            for city in cities
        }

    def _iter_settlements(self, path, sheet_name):
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if sheet_name not in workbook.sheetnames:
                raise CommandError(f"Workbook sheet does not exist: {sheet_name}")
            worksheet = workbook[sheet_name]
            rows = worksheet.iter_rows(values_only=True)
            try:
                headers = next(rows)
            except StopIteration as exc:
                raise CommandError(f"Workbook sheet is empty: {sheet_name}") from exc

            header_indexes = {
                self._normalize_header(value): index
                for index, value in enumerate(headers)
            }
            self._validate_headers(header_indexes)

            for row in rows:
                yield {
                    "object_level": self._cell(row, header_indexes, "object_level"),
                    "object_name": self._cell(row, header_indexes, "object_name"),
                    "region": self._cell(row, header_indexes, "region"),
                    "settlement": self._cell(row, header_indexes, "settlement"),
                    "latitude": self._cell(row, header_indexes, "latitude_dadata"),
                    "longitude": self._cell(row, header_indexes, "longitude_dadata"),
                }
        finally:
            workbook.close()

    def _validate_headers(self, header_indexes):
        required_headers = {
            "object_level",
            "object_name",
            "region",
            "settlement",
            "latitude_dadata",
            "longitude_dadata",
        }
        missing_headers = sorted(required_headers - set(header_indexes))
        if missing_headers:
            raise CommandError(
                "Workbook does not contain required columns: "
                + ", ".join(missing_headers)
            )

    def _import_settlement(
        self,
        settlement,
        regions_by_name,
        cities_by_key,
        country_code,
        dry_run,
        stats,
    ):
        if not self._is_settlement_level(settlement["object_level"]):
            stats["skipped_not_settlement"] += 1
            return

        name = self._normalize_name(settlement["settlement"] or settlement["object_name"])
        if not name:
            stats["skipped_without_name"] += 1
            return

        region = regions_by_name.get(self._normalize_key(settlement["region"]))
        if region is None:
            stats["skipped_without_region"] += 1
            return

        latitude = self._coordinate(settlement["latitude"])
        longitude = self._coordinate(settlement["longitude"])
        if latitude is False or longitude is False:
            stats["skipped_invalid_coordinates"] += 1
            return

        seen_key = (self._normalize_key(name), region.pk, country_code)
        if seen_key in self._seen_city_keys:
            stats["skipped_duplicate_in_file"] += 1
            return
        self._seen_city_keys.add(seen_key)

        existing = cities_by_key.get(seen_key)
        if dry_run:
            if (
                existing
                and existing.latitude == latitude
                and existing.longitude == longitude
            ):
                stats["unchanged"] += 1
            else:
                stats["updated" if existing else "created"] += 1
            return

        if existing:
            if existing.latitude == latitude and existing.longitude == longitude:
                stats["unchanged"] += 1
                return
            existing.latitude = latitude
            existing.longitude = longitude
            existing.save(update_fields=("latitude", "longitude", "updated_at"))
            stats["updated"] += 1
            return

        city = City.objects.create(
            name=name,
            region=region,
            country_code=country_code,
            latitude=latitude,
            longitude=longitude,
        )
        cities_by_key[seen_key] = city
        stats["created"] += 1

    def _cell(self, row, header_indexes, name):
        index = header_indexes[name]
        if index >= len(row):
            return ""
        value = row[index]
        return "" if value is None else value

    def _normalize_header(self, value):
        return str(value or "").strip()

    def _normalize_name(self, value):
        if value is None:
            return ""
        return str(value).strip()[:150]

    def _normalize_key(self, value):
        if value is None:
            return ""
        return " ".join(str(value).strip().casefold().replace("ё", "е").split())

    def _is_settlement_level(self, value):
        return self._normalize_key(value) in SETTLEMENT_LEVELS

    def _coordinate(self, value):
        if value in (None, ""):
            return None
        try:
            coordinate = Decimal(str(value)).quantize(
                COORDINATE_QUANT,
                rounding=ROUND_HALF_UP,
            )
        except (InvalidOperation, ValueError):
            return False
        return coordinate
